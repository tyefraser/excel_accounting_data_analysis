import pandas as pd
import os
import logging
import logging.config
from datetime import datetime
from openpyxl import Workbook
from contextlib import contextmanager

from utils import absolute_path
from utils import assert_file_extension

# Create a logger variable
logger = logging.getLogger(__name__)

def read_sheets_as_df(xls):
    logger.debug('\n\n\nRunning: read_sheets_as_df')

    # Get Sheet Names
    sheet_names = xls.sheet_names

    # Read each sheet into a DataFrame
    account_list_dfs_dict = {}
    for sheet_name in sheet_names:
        account_list_dfs_dict[sheet_name] = xls.parse(sheet_name)

    return account_list_dfs_dict

def basic_checks_acc_gen(
        account_list_dfs_dict
):
    logger.debug('\n\n\nRunning: basic_checks_acc_gen')

    # Pass unless fails
    pass_checks=True

    for account_listing in account_list_dfs_dict.keys():
        # All dfs must only have 1 column
        if not (account_list_dfs_dict[account_listing].shape[1] == 1):
            logger.info(f'ERROR: account_listing:{account_listing} contains more than one column, please review.')
            pass_checks=False

        # All values in the df must be unique
        if not (len(account_list_dfs_dict[account_listing].iloc[:, 0].unique()) == len(account_list_dfs_dict[account_listing].iloc[:, 0])):
            logger.info(f'ERROR: account_listing:{account_listing} contains duplicate values, please review.')
            pass_checks=False

    if pass_checks:
        logger.debug("All checks have passes.")
        return True
    else:
        logger.info('ERROR - please review above errors.')
        return False

def get_list_of_unique_accounts(
        account_list_dfs_dict
):
    logger.debug('\n\n\nRunning: get_list_of_unique_accounts')

    # Convert to lists
    account_lists_dict = {}
    all_accounts = []
    for account_listing in account_list_dfs_dict.keys():
        new_list=list(account_list_dfs_dict[account_listing].iloc[:,0])
        account_lists_dict[account_listing]=new_list
        all_accounts=all_accounts+new_list

    # Create list of all account with unique values only
    unique_accounts_list= list(set(all_accounts))

    logger.debug(f'account_lists_dict:{account_lists_dict}')
    logger.debug(f'list of all unique accounts:{unique_accounts_list}')
    return account_lists_dict, unique_accounts_list

def generate_check_account_order_dict(
        account_lists_dict,
        unique_accounts_list
):
    logger.debug('\n\n\nRunning: generate_check_account_order_dict')

    # Make a dictionary that details all accounts expected before and after another account
    check_account_order_dict = {}
    for account in unique_accounts_list:
        check_account_order_dict[account] = {
            'before': [],
            'after': [],
        }

    for account_listing in account_lists_dict.keys():
        logger.debug(f'Current account listing:{account_listing}')
        
        # Get the current account listing
        current_list=account_lists_dict[account_listing]

        # Get each account in the current account listing
        for account in current_list:
            # Find the index of the specific value
            index_of_account = current_list.index(account)

            # Get all values before the specific value
            accounts_before = current_list[:index_of_account]

            # Update before list
            check_account_order_dict[account]['before'] = list(set(
                check_account_order_dict[account]['before'] +
                accounts_before,
            ))

            # Get all values after the specific value
            accounts_after = current_list[index_of_account + 1:]

            # Update after list
            check_account_order_dict[account]['after'] = list(set(
                check_account_order_dict[account]['after'] +
                accounts_after
            ))

    logger.debug(f'check_account_order_dict:{check_account_order_dict}')
    return check_account_order_dict

def exclusive_list_check(list_1, list_2, detail):
    logger.debug('\n\n\nRunning: exclusive_list_check')

    # Passes unless it fails
    pass_checks=True
    logger.debug(f'{detail} listing (List 1):{list_1}')
    logger.debug(f'Should not be in (List 2):{list_2}')

    # Each item in list 1 should not be in list 2
    for item in list_1:
        if item in list_2:
            logger.info(f'{item} is incorrectly {detail} these values {list_2}')
            pass_checks = False

    return pass_checks

def account_order_checker(
        check_account_order_dict,
        account_list,
):
    logger.debug('\n\n\nRunning: account_order_checker')

    for account in account_list:
        logger.debug(f'checking:{account}')

        # Passes unless it fails
        pass_checks=True

        # Find the index of the specific value
        index_of_account = account_list.index(account)

        # Get all values before the specific value
        accounts_before = account_list[:index_of_account]
        logger.debug(f'\nBefore list: {accounts_before}')

        # Make sure values in before list shouldnt be in the after list
        if not exclusive_list_check(list_1 = accounts_before, list_2 = check_account_order_dict[account]['after'], detail = 'before'):
                logger.info('FAIL: An account before this account ({account}) in the list should not be before it.')
                pass_checks=False

        # Get all values after the specific value
        accounts_after = account_list[index_of_account + 1:]
        logger.debug(f'\nAfter list: {accounts_after}')

        # Make sure values in the after lsit shouldnt be in the before list
        if not exclusive_list_check(list_1 = accounts_after, list_2 = check_account_order_dict[account]['before'], detail = 'after'):
                logger.info('FAIL: An account after this account ({account}) in the list should not be after it.')
                pass_checks=True
        
        if not pass_checks:
                logger.info('This list fails the checks\n')

        return(pass_checks)

def complete_account_order_checker(
        check_account_order_dict,
        account_list_lists,
):
    logger.debug('\n\n\nRunning: complete_account_order_checker')

    # Passes unless fails
    pass_checks = True

    for account_list in account_list_lists:
        logger.debug(f'Checking list: {account_list}')
        if not account_order_checker(
            check_account_order_dict=check_account_order_dict,
            account_list=account_list_lists[account_list],
        ):
            pass_checks=False
    
    return pass_checks

def correctly_ordered_list(
        check_account_order_dict,
):
    logger.debug('\n\n\nRunning: correctly_ordered_list')

    ordered_account_list = []
    for accoount in check_account_order_dict.keys():
        logger.debug(f'\nAssessing:{accoount}')

        before_list = check_account_order_dict[accoount]['before']
        logger.debug(f'before_list:{before_list}')
        max_before_pos=None
        for item in before_list:
            if (item in ordered_account_list):
                pos=ordered_account_list.index(item)
                if max_before_pos == None:
                    max_before_pos = pos
                elif pos > max_before_pos:
                    max_before_pos = pos
        logger.debug(f'max_before_pos:{max_before_pos}')

        after_list = check_account_order_dict[accoount]['after']
        logger.debug(f'after_list:{after_list}')
        min_after_pos=None
        for item in after_list:
            if (item in ordered_account_list):
                pos=ordered_account_list.index(item)
                if min_after_pos == None:
                    min_after_pos = pos
                elif pos < min_after_pos:
                    min_after_pos=pos
        logger.debug(f'min_after_pos:{min_after_pos}')

        if (max_before_pos == None) and (min_after_pos == None):
            ordered_account_list = ordered_account_list + [accoount]
            logger.debug("Nones")
            logger.debug(ordered_account_list)

        elif max_before_pos == None:
            logger.debug(f'min_after_pos:{min_after_pos}')
            
            # Add after the last item it needs to be before
            ordered_account_list.insert(min_after_pos, accoount)
            logger.debug(ordered_account_list)

        else:
            logger.debug("No after position indicator")
            logger.debug(f'max_before_pos:{max_before_pos}')

            # max_before_pos is not None, insert one before max_before_pos
            ordered_account_list.insert(max_before_pos+1, accoount)
            logger.debug(ordered_account_list)
            
    return ordered_account_list



@contextmanager
def openpyxl_context_manager(file_path):
    wb = Workbook()
    try:
        yield wb
    finally:
        wb.save(file_path)

def document_account_lists(ordered_account_list, account_lists_dict):
    logger.debug('\n\n\nRunning: document_account_lists')

    # Specify the sheet title
    sheet_title = 'MyAccountsSheet'

    # Add in the ordered account list
    excel_file_name=f'Ordered_Accounts_{datetime.now().strftime("%Y_%m_%d_at_%I_%M%p_")}.xlsx'
    logger.debug(f'excel_file_name:{excel_file_name}')
    excel_file_path=absolute_path('results/' + excel_file_name)
    logger.debug(f'excel_file_path:{excel_file_path}')
    
    with openpyxl_context_manager(absolute_path(f'get_unique_accounts/results/Ordered_Accounts_{datetime.now().strftime("%Y_%m_%d_at_%I_%M%p_")}.xlsx')) as wb:
        # Delete the default 'Sheet' if it exists
        default_sheet = wb['Sheet']
        if default_sheet:
            wb.remove(default_sheet)

        # Activate the sheet by title or create a new one if it doesn't exist
        ws = wb[sheet_title] if sheet_title in wb.sheetnames else wb.create_sheet(title=sheet_title)

        ws.cell(row=1, column=1, value='Accounts List - Complete')
        for index, account in enumerate(ordered_account_list):
            ws.cell(row=index + 2, column=1, value=account)

        # Iterate over the account lists add accounts in the order and position of the ordered list.
        for list_index, account_list_name in enumerate(list(account_lists_dict.keys())):
            # Get column position for the current list
            col_pos = list_index + 2

            # Add the account name as the heading
            ws.cell(row=1, column=col_pos, value=account_list_name)

            # Add in the accounts through the order of the ordered_account_list
            current_account_list = account_lists_dict[account_list_name]
            for row_index, account in enumerate(ordered_account_list):
                if account in current_account_list:
                    ws.cell(row=row_index + 2, column=col_pos, value=account)

        logger.info(f'Excel saved as: {excel_file_name}')

def main(
        file_name: str = "account_lists_single_tst.xlsx",
):
    logger.debug('\n\n\nRunning: Main')

    # Check file is Excel
    assert_file_extension(file_name, expected_extension='.xlsx')

    # Read the Excel file
    xls = pd.ExcelFile(file_name)

    # Read in sheets as dfs
    account_list_dfs_dict = read_sheets_as_df(xls)
    
    # Run basic account list checks
    assert (basic_checks_acc_gen(account_list_dfs_dict=account_list_dfs_dict)), ('Basic checks have failed')

    # Get a list of unique accounts from all lists
    account_lists_dict, unique_accounts_list=get_list_of_unique_accounts(account_list_dfs_dict)

    # Create a dictionary to use as account ofer checker
    check_account_order_dict=generate_check_account_order_dict(account_lists_dict, unique_accounts_list)

    # Check all account lists are in the expected order
    assert complete_account_order_checker(
            check_account_order_dict,
            account_lists_dict,
    ), 'Some accounts are not in the correctr order, please review'

    # At this point you have a list of acount lists
    # These account lists are all in the correct order

    # Generate the list of all accounts in the correct order
    ordered_account_list = correctly_ordered_list(check_account_order_dict)

    # Generate Excel with details on accounts
    document_account_lists(ordered_account_list,account_lists_dict,)

    return ordered_account_list
